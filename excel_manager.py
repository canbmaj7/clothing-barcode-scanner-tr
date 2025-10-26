"""
Excel Manager - Ürün sonuçlarını Excel dosyasına kaydeder
"""
import os
from datetime import datetime
import xlsxwriter


class ExcelManager:
    """Excel dosyası yönetimi"""
    
    def __init__(self):
        """Excel manager başlatır"""
        self.excel_dir = os.path.join("outputs", "excel")
        self.ensure_directory()
        
        # Bugünkü tarih ile dosya adı
        today = datetime.now().strftime("%Y-%m-%d")
        self.filename = os.path.join(self.excel_dir, f"tum_urunler_{today}.xlsx")
        
        # Eğer dosya zaten varsa, ona append edeceğiz
        # xlsxwriter append desteklemiyor, openpyxl kullanmalıyız
        self.products = []
        
    def ensure_directory(self):
        """Excel klasörünü oluşturur"""
        if not os.path.exists(self.excel_dir):
            os.makedirs(self.excel_dir)
    
    def add_products(self, products_list, brand_name):
        """
        Ürünleri listeye ekler
        
        Args:
            products_list (list): [(barkod, ürün_adı), ...] formatında liste
            brand_name (str): Marka adı
        """
        for barkod, urun_adi in products_list:
            self.products.append({
                'barkod': barkod,
                'urun_adi': urun_adi,
                'brand': brand_name
            })
    
    def save_to_excel(self):
        """
        Tüm ürünleri Excel dosyasına kaydeder
        
        Returns:
            tuple: (başarılı_mı, hata_mesajı)
        """
        try:
            # Dosya açık mı kontrol et
            if self._is_file_open():
                return False, "Excel dosyası açık! Lütfen kapatıp tekrar deneyin."
            
            # Eğer dosya varsa, mevcut verileri oku
            existing_products = []
            if os.path.exists(self.filename):
                existing_products = self._read_existing_excel()
            
            # Mevcut + yeni verileri birleştir
            all_products = existing_products + self.products
            
            # Yeni Excel dosyası oluştur
            workbook = xlsxwriter.Workbook(self.filename)
            worksheet = workbook.add_worksheet('Ürünler')
            
            # Başlık formatı
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': 'white',
                'border': 1
            })
            
            # Başlıkları yaz
            worksheet.write('A1', 'Barkod', header_format)
            worksheet.write('B1', 'Ürün Adı', header_format)
            
            # Sütun genişlikleri
            worksheet.set_column('A:A', 20)
            worksheet.set_column('B:B', 60)
            
            # Verileri yaz
            for i, product in enumerate(all_products, start=2):
                worksheet.write(f'A{i}', product['barkod'])
                # Marka adını ürün adının başına ekle
                full_product_name = f"{product['brand']} {product['urun_adi']}"
                worksheet.write(f'B{i}', full_product_name)
            
            workbook.close()
            
            # Listeyi temizle
            self.products = []
            
            return True, None
            
        except Exception as e:
            return False, f"Excel yazma hatası: {e}"
    
    def _is_file_open(self):
        """
        Excel dosyasının açık olup olmadığını kontrol eder
        
        Returns:
            bool: Dosya açıksa True
        """
        if not os.path.exists(self.filename):
            return False
        
        try:
            # Dosyayı yazma modunda açmayı dene
            with open(self.filename, 'a'):
                pass
            return False
        except IOError:
            return True
    
    def _read_existing_excel(self):
        """
        Mevcut Excel dosyasını okur (openpyxl ile)
        
        Returns:
            list: Mevcut ürünler listesi
        """
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(self.filename)
            worksheet = workbook.active
            
            products = []
            # İlk satır başlık, 2. satırdan başla
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:  # Barkod ve ürün adı varsa
                    # Mevcut Excel'de marka adı ürün adının içinde olabilir
                    # Eğer öyleyse, ayrıştırmaya çalış
                    product_name = str(row[1])
                    brand = "BİLİNMEYEN"  # Varsayılan
                    
                    # Marka adlarını kontrol et
                    for possible_brand in ["BERSHKA", "H&M", "ZARA", "MANGO", "MAVİ"]:
                        if product_name.startswith(possible_brand + " "):
                            brand = possible_brand
                            product_name = product_name[len(possible_brand)+1:]  # Marka adını çıkar
                            break
                    
                    products.append({
                        'barkod': str(row[0]),
                        'urun_adi': product_name,
                        'brand': brand
                    })
            
            workbook.close()
            return products
            
        except Exception as e:
            print(f"⚠️  Mevcut Excel okunamadı: {e}")
            return []


def parse_product_line(line, brand_name):
    """
    TXT dosyasındaki ürün satırını parse eder
    
    Args:
        line (str): "barkod MARKA Ürün Adı" formatında satır
        brand_name (str): Marka adı (BERSHKA, H&M, vb.)
    
    Returns:
        tuple: (barkod, ürün_adı) veya None
    """
    parts = line.split(f" {brand_name} ", 1)
    if len(parts) == 2:
        barkod = parts[0].strip()
        urun_adi = parts[1].strip()
        return (barkod, urun_adi)
    return None

